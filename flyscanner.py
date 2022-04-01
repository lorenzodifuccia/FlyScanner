import os
import datetime
from typing import Union

import providers


class FlyScanner:
    def __init__(self, provider, print_all=False, print_detail=True):
        self.provider = providers.PROVIDERS[provider]()
        self.searching = False

        self.print_all = print_all
        self.print_detail = print_detail

        self.save = False
        self.workbook = None
        self.sheet = None

    def prepare_location(self, iata):
        return self.provider.prepare_location({"iata": iata})

    def search(self, adults: int, date: str, departure_iata: dict, destination_iata: dict):
        search_resp = self.provider.search(adults, date, departure_iata, destination_iata)

        for solution in search_resp["result"]:
            if self.save:
                self.save_to_file(solution)

            print("  \u2022 ✈️  FOUND:", solution["price"], solution["price_currency"], "-",
                  solution["carrier"], "-", "🕓", solution["duration"], "✈️ ")

            if not self.print_detail:
                break

            print("  Departure:", solution["departure_date"])
            print("  From:", solution["departure_location"])
            print("  Arrival:", solution["arrival_date"])
            print("  To:", solution["arrival_location"])

            if solution["stops"] > 1:
                print("  Stops:", solution["stops"] - 1, "-", "No Fly Duration:", solution["stops_duration"])

                for stop in range(solution["stops"]):
                    print("   \u2022 ✈️  STOP %d  ✈️ " % (stop + 1))
                    print("   Departure:", solution["stops_detail"][stop]["departure_date"])
                    print("   From:", solution["stops_detail"][stop]["departure_location"])
                    print("   Arrival:", solution["stops_detail"][stop]["arrival_date"])
                    print("   To:", solution["stops_detail"][stop]["arrival_location"])
                    print("   Duration:", solution["stops_detail"][stop]["duration"])

            if len(solution["discounts"]) > 0:
                print("  Available Discounts:")
                for discount in solution["discounts"]:
                    print("   Discount:", discount["price"], discount["price_currency"], discount["reason"])

            print()

            if not self.print_all:
                break

    def start_search(self, departure_iata, destination_iata, from_date, to_date=None, adults=1,
                     save: Union[bool, str] = False):

        searching_date = datetime.date.fromisoformat(from_date)
        stop_date = datetime.date.fromisoformat(to_date) if to_date else None

        if save:
            self.save = os.path.expanduser(save)
            if os.path.exists(self.save):
                self.workbook = openpyxl.load_workbook(self.save)
                self.sheet = self.workbook.create_sheet()

            else:
                self.workbook = openpyxl.Workbook()
                self.sheet = self.workbook.active

            self.sheet.title = "%s %s (%s)" % (departure_iata["iata"], destination_iata["iata"],
                                               datetime.datetime.now().strftime("%y-%m-%d %H.%M.%S"))

        self.searching = True
        print("Provider: %s" % self.provider.NAME)
        while self.searching:
            print("Searching for %s ..." % searching_date)
            self.search(adults, searching_date.isoformat(), departure_iata, destination_iata)

            if searching_date == stop_date:
                self.searching = False
                exit()

            searching_date = searching_date + datetime.timedelta(1)

    def stop_search(self):
        self.searching = False

        if self.save:
            self.workbook.save(filename=self.save)

    def save_to_file(self, solution: dict):
        if self.sheet.max_column > 1:
            keys = list(x[0].value for x in self.sheet.iter_cols(max_row=1))
            remaining_keys = list(set(keys) ^ set(solution.keys()))

        else:
            keys = providers.FLIGHT_DEFAULT + list(set(solution.keys()) ^ set(providers.FLIGHT_DEFAULT))
            remaining_keys = []
            self.sheet.append(keys)

        row_data = []
        for k in keys + remaining_keys:
            v = solution.get(k, "")
            row_data.append("\n".join(str(x) for x in v) if isinstance(v, (list, set, tuple)) else
                            (str(v) if isinstance(v, dict) else v))

        self.sheet.append(row_data)


if __name__ == "__main__":
    import time
    import argparse
    import threading

    parser = argparse.ArgumentParser()

    mutual_mode = parser.add_mutually_exclusive_group(required=True)

    mutual_mode.add_argument("--search", action='store_true', help="Search on the selected provider")
    mutual_mode.add_argument("--autocomplete", nargs="+", help="Helper to find IATA of city and airport")
    mutual_mode.add_argument("--providers", action='store_true', help="List available providers")

    search_group = parser.add_argument_group("search options")
    search_group.add_argument("--provider", default="eDreams", choices=providers.PROVIDERS.keys())

    search_group.add_argument("--departure", "--from", "-f", type=str, metavar="IATA")
    search_group.add_argument("--destination", "--to", "-t", type=str, metavar="IATA")
    search_group.add_argument("--date", "-d", type=str, metavar="YYYY-MM-DD")
    search_group.add_argument("--to-date", type=str, metavar="YYYY-MM-DD")
    search_group.add_argument("--passengers", metavar="#", default=1, type=int)

    output_group = parser.add_argument_group("output options")
    output_group_mutual = output_group.add_mutually_exclusive_group()
    output_group_mutual.add_argument("--all", action='store_true',
                                     help="Show all fly options for the day, not only the cheepest")
    output_group_mutual.add_argument("--list", action='store_true', help="Reduce verbosity")

    output_group.add_argument("--save", nargs='?', type=str, default=False, const="~/Desktop/FlyScannerTrips.xlsx")

    args = parser.parse_args()
    print("\n"
          "    ________      _____                                 \n"
          "   / ____/ /_  __/ ___/_________ _____  ____  ___  _____\n"
          "  / /_  / / / / /\__ \/ ___/ __ `/ __ \/ __ \/ _ \/ ___/\n"
          " / __/ / / /_/ /___/ / /__/ /_/ / / / / / / /  __/ /    \n"
          "/_/   /_/\__, //____/\___/\__,_/_/ /_/_/ /_/\___/_/     \n"
          "        /____/                                          \n")

    if args.autocomplete:
        for suggestion_obj in providers.PROVIDERS[args.provider].autocomplete(" ".join(args.autocomplete)):
            providers.PROVIDERS[args.provider].print_autocomplete(suggestion_obj)
            print()

    elif args.providers:
        print("Providers:\n  \u2022", "\n  \u2022 ".join(providers.PROVIDERS.keys()))

    elif args.search:
        if not all((args.departure, args.destination, args.date)):
            parser.error("--departure, --destination, --date required!")

        elif len(args.departure) != 3 or len(args.destination) != 3:
            parser.error("--departure and --destination must be IATA code (use --autocomplete to find them)")

        if args.save:
            import openpyxl

        fly = FlyScanner(args.provider, args.all, not args.list)

        departure = fly.prepare_location(args.departure)
        if not departure:
            parser.error("--departure: invalid IATA " + args.departure)

        destination = fly.prepare_location(args.destination)
        if not destination:
            parser.error("--destination: invalid IATA " + args.destination)

        searching_th = threading.Thread(name="FlySearching",
                                        target=fly.start_search,
                                        kwargs={
                                            "departure_iata": departure,
                                            "destination_iata": destination,
                                            "from_date": args.date,
                                            "to_date": args.to_date,
                                            "adults": args.passengers,
                                            "save":  args.save
                                        })

        searching_th.start()
        while searching_th.is_alive():
            try:
                time.sleep(1)

            except KeyboardInterrupt:
                break

        fly.stop_search()

        try:
            searching_th.join()

        except KeyboardInterrupt:
            pass

    print("Bye!")
